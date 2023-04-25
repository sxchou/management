from django.shortcuts import render, redirect
from openpyxl import load_workbook
from app01 import models
from app01.utils.pagination import Pagination


def department_list(request):
    """部门列表"""
    # if not request.session.get('info'):
    #     return redirect('/login/')  # 验证用户是否登录（每个页面都要添加，过于繁琐，使用中间件解决）
    data_dict = {}
    query = request.GET.get('query', '')
    if query:
        data_dict['department_name__contains'] = query
    departments = models.Department.objects.filter(**data_dict)
    page_obj = Pagination(request, departments)
    context = {
        'query': query,
        'department_list': page_obj.page_queryset,
        'page_string': page_obj.html()
    }
    return render(request, 'department_list.html', context)


def department_add(request):
    """新建部门"""
    if request.method == 'GET':
        return render(request, 'department_add.html')
    department_name = request.POST.get('department_name')
    models.Department.objects.create(department_name=department_name)
    return redirect('/department/list/')


def department_delete(request, index_id):
    """删除部门"""
    models.Department.objects.filter(id=index_id).delete()
    return redirect('/department/list/')


def department_edit(request, index_id):
    """修改部门"""
    if request.method == 'GET':
        row_obj = models.Department.objects.filter(id=index_id).first()
        return render(request, 'department_edit.html', {'row_obj': row_obj})
    new_department_name = request.POST.get('department_name')
    models.Department.objects.filter(id=index_id).update(department_name=new_department_name)
    return redirect('/department/list/')


def department_multi(request):
    file_obj = request.FILES.get('excel_file')
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        exists = models.Department.objects.filter(department_name=row[0].value).exists()
        if not exists:
            models.Department.objects.create(department_name=row[0].value)
    return redirect('/department/list/')

